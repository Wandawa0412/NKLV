"""Excel engine — generates Excel files from WorkLog data using the official template."""
import os
import copy
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from core.models import WorkLog, WorkItem
from core.app_paths import TEMPLATE_PATH, OUTPUT_DIR
from core.date_utils import format_display_date, parse_date

# ── Style constants ──────────────────────────────────────────
FONT_TITLE = Font(name="Times New Roman", size=14, bold=True)
FONT_HEADER = Font(name="Times New Roman", size=12, bold=True)
FONT_DATA = Font(name="Times New Roman", size=12)
FONT_TOTAL = Font(name="Times New Roman", size=12, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

NUMBER_FORMAT = "#,##0"
NUM_COLS = 5  # Ngày, Nội dung, SL, Đơn giá, Thành tiền

CHARS_PER_LINE = 42
LINE_HEIGHT = 15.0


def _ordered_export_items(items: list[WorkItem]) -> list[WorkItem]:
    """Preserve UI order so continuation rows stay under the correct date row."""
    return list(items)


class ExcelEngine:
    def __init__(self, template_path: str = TEMPLATE_PATH):
        self.template_path = template_path
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def _estimate_row_height(self, content: str) -> float:
        if not content:
            return LINE_HEIGHT
        lines = content.count("\n") + 1
        for line in content.split("\n"):
            if len(line) > CHARS_PER_LINE:
                lines += (len(line) - 1) // CHARS_PER_LINE
        return max(LINE_HEIGHT, lines * LINE_HEIGHT)

    def _create_workbook_from_template(self, log: WorkLog):
        wb = load_workbook(self.template_path)
        ws = wb.active

        for merge in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merge))

        # Clear ALL data rows — use generous range to handle reused templates
        clear_rows = max(ws.max_row + 10, 500)
        for row in range(1, clear_rows + 1):
            for col in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.font = FONT_DATA
                cell.border = Border()
                cell.fill = PatternFill()
                cell.alignment = Alignment()
                cell.number_format = "General"

        # Row 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        ws.cell(row=1, column=1, value="NHẬT KÝ LÀM VIỆC").font = FONT_TITLE
        ws["A1"].alignment = ALIGN_CENTER
        ws.row_dimensions[1].height = 25

        # Row 2: Customer
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
        ws.cell(row=2, column=1, value=f"Khách hàng: {log.customer_name}").font = FONT_HEADER
        ws["A2"].alignment = ALIGN_LEFT
        ws.cell(row=2, column=4, value=f"Ngày phiếu: {format_display_date(log.work_date)}").font = FONT_HEADER
        ws["D2"].alignment = ALIGN_CENTER
        ws.row_dimensions[2].height = 22

        # Row 3: Headers — A:Ngày  B:Nội dung  C:SL  D:Đơn giá  E:Thành tiền
        headers = [
            ("Ngày", ALIGN_CENTER),
            ("Nội dung", ALIGN_CENTER_WRAP),
            ("Số\nlượng", ALIGN_CENTER_WRAP),
            ("Đơn giá", ALIGN_CENTER),
            ("Thành tiền", ALIGN_CENTER),
        ]
        for col_idx, (text, align) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = FONT_HEADER
            cell.alignment = align
            cell.border = THIN_BORDER
            cell.fill = HEADER_FILL
        ws.row_dimensions[3].height = 30

        # Preserve grouped row order from the UI/database instead of flattening by date.
        sorted_items = _ordered_export_items(log.items)

        # Data rows
        data_start = 4
        for i, item in enumerate(sorted_items):
            row_num = data_start + i

            # A: Ngày
            ws.cell(row=row_num, column=1, value=item.date).font = FONT_DATA
            ws.cell(row=row_num, column=1).alignment = ALIGN_CENTER
            ws.cell(row=row_num, column=1).border = THIN_BORDER

            # B: Nội dung
            ws.cell(row=row_num, column=2, value=item.content).font = FONT_DATA
            ws.cell(row=row_num, column=2).alignment = ALIGN_LEFT_WRAP
            ws.cell(row=row_num, column=2).border = THIN_BORDER

            # C: Số lượng
            ws.cell(row=row_num, column=3, value=item.quantity).font = FONT_DATA
            ws.cell(row=row_num, column=3).alignment = ALIGN_CENTER
            ws.cell(row=row_num, column=3).border = THIN_BORDER

            # D: Đơn giá
            ws.cell(row=row_num, column=4, value=item.unit_price).font = FONT_DATA
            ws.cell(row=row_num, column=4).alignment = ALIGN_RIGHT
            ws.cell(row=row_num, column=4).border = THIN_BORDER
            ws.cell(row=row_num, column=4).number_format = NUMBER_FORMAT

            # E: Thành tiền = C * D
            ws.cell(row=row_num, column=5, value=f"=C{row_num}*D{row_num}").font = FONT_DATA
            ws.cell(row=row_num, column=5).alignment = ALIGN_RIGHT
            ws.cell(row=row_num, column=5).border = THIN_BORDER
            ws.cell(row=row_num, column=5).number_format = NUMBER_FORMAT

            ws.row_dimensions[row_num].height = max(18, self._estimate_row_height(item.content))

        # Total row
        total_row = data_start + len(sorted_items)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = FONT_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=1).border = THIN_BORDER
        ws.cell(row=total_row, column=1).fill = TOTAL_FILL
        for col in range(2, 5):
            ws.cell(row=total_row, column=col).border = THIN_BORDER
            ws.cell(row=total_row, column=col).fill = TOTAL_FILL

        ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{total_row-1})").font = FONT_TOTAL
        ws.cell(row=total_row, column=5).alignment = ALIGN_RIGHT
        ws.cell(row=total_row, column=5).border = THIN_BORDER
        ws.cell(row=total_row, column=5).fill = TOTAL_FILL
        ws.cell(row=total_row, column=5).number_format = NUMBER_FORMAT
        ws.row_dimensions[total_row].height = 22

        # Column widths
        ws.column_dimensions["A"].width = 13    # Ngày
        ws.column_dimensions["B"].width = 40    # Nội dung
        ws.column_dimensions["C"].width = 8     # SL
        ws.column_dimensions["D"].width = 13    # Đơn giá
        ws.column_dimensions["E"].width = 14    # Thành tiền

        # Print
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        return wb

    def export_single(self, log: WorkLog, output_path: str | None = None) -> str:
        wb = self._create_workbook_from_template(log)
        if not output_path:
            safe_name = log.customer_name.replace("/", "-").replace("\\", "-").replace(":", "-").strip()
            output_path = os.path.join(
                OUTPUT_DIR,
                f"NKLV_{safe_name}_{log.work_date.strftime('%Y%m%d')}.xlsx",
            )
        try:
            wb.save(output_path)
        except PermissionError:
            raise PermissionError(
                f"Không thể ghi file (có thể đang mở):\n{output_path}"
            )
        return output_path

    def export_batch(self, logs: list[WorkLog], output_dir: str | None = None) -> list[str]:
        output_dir = output_dir or OUTPUT_DIR
        os.makedirs(output_dir, exist_ok=True)
        paths = []
        used_names: dict[str, int] = {}  # Track duplicate filenames
        for log in logs:
            safe_name = log.customer_name.replace("/", "-").replace("\\", "-").replace(":", "-").strip()
            base_name = f"NKLV_{safe_name}_{log.work_date.strftime('%Y%m%d')}"
            # Handle duplicate names with counter suffix
            if base_name in used_names:
                used_names[base_name] += 1
                file_name = f"{base_name}_{used_names[base_name]}.xlsx"
            else:
                used_names[base_name] = 0
                file_name = f"{base_name}.xlsx"
            file_path = os.path.join(output_dir, file_name)
            paths.append(self.export_single(log, output_path=file_path))
        return paths

    def export_multi_sheet(self, logs: list[WorkLog], output_path: str | None = None) -> str:
        if not output_path:
            output_path = os.path.join(OUTPUT_DIR, f"NKLV_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        from openpyxl import Workbook
        combined_wb = Workbook()
        combined_wb.remove(combined_wb.active)

        for i, log in enumerate(logs):
            wb = self._create_workbook_from_template(log)
            source_ws = wb.active
            safe_name = log.customer_name[:25].replace("/", "-") if log.customer_name else f"Sheet{i+1}"
            sheet_name = f"{i+1}.{safe_name}"[:31]
            # Handle duplicate sheet names
            existing_names = [ws.title for ws in combined_wb.worksheets]
            if sheet_name in existing_names:
                suffix = 2
                while f"{sheet_name[:28]}_{suffix}"[:31] in existing_names:
                    suffix += 1
                sheet_name = f"{sheet_name[:28]}_{suffix}"[:31]
            target_ws = combined_wb.create_sheet(title=sheet_name)

            for row in source_ws.iter_rows():
                for cell in row:
                    tc = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        tc.font = copy.copy(cell.font)
                        tc.border = copy.copy(cell.border)
                        tc.fill = copy.copy(cell.fill)
                        tc.number_format = cell.number_format
                        tc.protection = copy.copy(cell.protection)
                        tc.alignment = copy.copy(cell.alignment)

            for merge in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merge))
            for col_letter, dim in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_letter].width = dim.width
            for row_num, dim in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_num].height = dim.height

        try:
            combined_wb.save(output_path)
        except PermissionError:
            raise PermissionError(
                f"Không thể ghi file (có thể đang mở):\n{output_path}"
            )
        return output_path

    def import_from_excel(self, file_path: str) -> WorkLog:
        """Import a work log from an existing Excel file.

        Handles real-world formats:
          - Date column can be datetime objects or strings
          - Price column may be empty
          - Total row may not have "Tổng" text
          - Supports 5-col and 6-col layouts
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Parse metadata row
        customer_raw = str(ws.cell(row=2, column=1).value or "").strip()
        customer_name = customer_raw
        if customer_raw.lower().startswith(("khách hàng:", "khach hang:")):
            customer_name = customer_raw.split(":", 1)[1].strip()

        work_date_raw = (
            str(ws.cell(row=2, column=4).value or "").strip()
            or str(ws.cell(row=2, column=5).value or "").strip()
        )
        if work_date_raw.lower().startswith(("ngày phiếu:", "ngay phieu:")):
            work_date_raw = work_date_raw.split(":", 1)[1].strip()
        parsed_work_date = parse_date(work_date_raw)

        # Detect column layout by reading row 3 headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=3, column=col).value or "").strip().replace("\n", " ").lower()
            headers.append(val)

        # Smart column mapping by header content
        col_date = col_content = col_qty = col_price = col_total = None
        for i, h in enumerate(headers):
            col_1based = i + 1
            if "ngày" in h or "ngay" in h:
                col_date = col_1based
            elif "nội dung" in h or "noi dung" in h or "content" in h:
                col_content = col_1based
            elif ("số" in h and "lượng" in h) or "so luong" in h or h == "sl":
                col_qty = col_1based
            elif "đơn giá" in h or "don gia" in h:
                col_price = col_1based
            elif "thành tiền" in h or "thanh tien" in h:
                col_total = col_1based

        # Fallback: assume standard 5-col layout if detection fails
        if col_date is None:
            col_date = 1
        if col_content is None:
            col_content = 2
        if col_qty is None:
            col_qty = 3
        if col_price is None and col_total is not None:
            # No separate price column — use total column for price (qty=1 assumed)
            col_price = col_total
        elif col_price is None:
            col_price = 4

        # Parse data rows (row 4 onwards)
        items = []
        for row in range(4, ws.max_row + 1):
            # Read raw values
            date_raw = ws.cell(row=row, column=col_date).value
            content_raw = ws.cell(row=row, column=col_content).value

            # Detect total/empty row — stop parsing
            first_cell_str = str(date_raw or "").strip().lower()
            if "tổng" in first_cell_str or "tong" in first_cell_str:
                break

            content = str(content_raw or "").strip()
            if not content and not date_raw:
                # Truly empty row or total row without label
                break

            # ── Parse date ──
            date_str = ""
            if date_raw is not None:
                if hasattr(date_raw, 'strftime'):
                    # datetime object → convert to dd/MM/yyyy
                    date_str = date_raw.strftime("%d/%m/%Y")
                else:
                    date_str = str(date_raw).strip()
                    # Try to detect and normalize common date formats
                    # "2026-01-12" → "12/01/2026"
                    if "-" in date_str and len(date_str) == 10:
                        try:
                            dt = datetime.strptime(date_str, "%Y-%m-%d")
                            date_str = dt.strftime("%d/%m/%Y")
                        except ValueError:
                            pass

            # ── Parse quantity ──
            qty_raw = ws.cell(row=row, column=col_qty).value
            try:
                qty = int(float(str(qty_raw).replace(",", ""))) if qty_raw else 1
            except (ValueError, TypeError):
                qty = 1

            # ── Parse price ──
            price_raw = ws.cell(row=row, column=col_price).value
            try:
                price = float(str(price_raw).replace(",", "")) if price_raw else 0.0
            except (ValueError, TypeError):
                price = 0.0

            items.append(WorkItem(
                date=date_str,
                content=content,
                quantity=max(0, qty),
                unit_price=max(0.0, price),
            ))

        derived_work_date = parsed_work_date or self._derive_work_date(items) or date.today()

        log = WorkLog(
            customer_name=customer_name if customer_name else "Không rõ",
            items=items,
            work_date=derived_work_date,
            created_at=datetime.now(),
        )
        return log

    @staticmethod
    def _derive_work_date(items: list[WorkItem]) -> date | None:
        parsed_dates = [
            parsed
            for item in items
            if (parsed := parse_date(item.date)) is not None
        ]
        if not parsed_dates:
            return None
        return min(parsed_dates)

